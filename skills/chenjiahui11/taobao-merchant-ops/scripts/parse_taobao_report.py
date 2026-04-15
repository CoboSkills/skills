"""
淘宝生意参谋报表解析器
按 mapping.final.json 规则解析Excel输出结构化JSON
"""
import openpyxl
import json
import os
import sys
from datetime import datetime

# Force UTF-8
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

MAPPING_PATH = r"C:\Users\22066\Desktop\mapping.final.json"
EXCEL_DIR = r"C:\Users\22066\Desktop\2.0\taobao_tool\scripts\downloads"
OUTPUT_DIR = r"C:\Users\22066\Desktop\openclawskills\openclaw-web-extract\output"


def load_mapping(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def find_header_index(ws, target_names):
    """在header行中查找目标列名，返回 {target_name: col_index}"""
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_map = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        # 尝试GBK解码（生意参谋Excel编码）
        try:
            decoded = cell.encode("latin1").decode("gbk")
        except Exception:
            decoded = str(cell)
        header_map[decoded] = idx
    # 按target_names顺序返回
    result = {}
    for name in target_names:
        result[name] = header_map.get(name, None)
    return result, header_map


def get_row1_data(ws, header_map):
    """获取第1行数据（实际数据行）"""
    data_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    result = {}
    for col_name, col_idx in header_map.items():
        if col_idx is not None and col_idx < len(data_row):
            result[col_name] = data_row[col_idx]
    return result


def safe_number(value):
    """统一数值类型：去千分位，处理None"""
    if value is None or str(value).strip() == "" or str(value).upper() == "NULL":
        return None
    s = str(value).replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return str(value)


def compute_field(expression, data):
    """计算派生字段，处理除零场景"""
    try:
        # 简单eval，只支持 + - * / 和括号
        # 把字段名替换成实际值
        expr = expression
        for field, value in data.items():
            if value is not None:
                expr = expr.replace(field, str(value))
            else:
                return None
        result = eval(expr)
        if isinstance(result, float) and (result != result or abs(result) == float('inf')):
            return None  # NaN or Inf
        return round(result, 4) if result else result
    except Exception:
        return None


def parse_excel(excel_path, mapping):
    """解析单个Excel文件"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # 构建完整header索引
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    full_header_map = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        try:
            decoded = cell.encode("latin1").decode("gbk")
        except Exception:
            decoded = str(cell)
        full_header_map[decoded] = idx

    # 获取数据行
    data_row = None
    for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
        data_row = row
        break

    def get_val(col_name):
        idx = full_header_map.get(col_name)
        if idx is None or idx >= len(data_row):
            return None
        val = data_row[idx]
        return safe_number(val)

    # 构建原始数据字典
    raw_data = {col: get_val(col) for col in full_header_map.keys()}

    # 应用直接映射
    result = {}
    for out_key, in_key in mapping["direct_mapping"].items():
        result[out_key] = raw_data.get(in_key)

    # 应用重命名映射
    for out_key, in_key in mapping["renamed_mapping"].items():
        result[out_key] = raw_data.get(in_key)

    # 应用计算映射
    for out_key, config in mapping["computed_mapping"].items():
        expr = config["expression"]
        deps = config["depends_on"]
        # 先替换依赖字段
        eval_data = {k: raw_data.get(k) for k in deps}
        result[out_key] = compute_field(expr, eval_data)

    # 数值字段统一处理
    for k, v in result.items():
        result[k] = safe_number(v) if v is not None else None

    return result, raw_data


def main():
    # 加载映射规则
    mapping = load_mapping(MAPPING_PATH)

    # 查找最新的Excel
    files = [f for f in os.listdir(EXCEL_DIR) if f.endswith(".xlsx")]
    if not files:
        print("ERROR: No Excel files found")
        return

    latest = max(files, key=lambda f: os.path.getmtime(os.path.join(EXCEL_DIR, f)))
    excel_path = os.path.join(EXCEL_DIR, latest)
    print(f"Parsing: {latest}")

    # 解析
    parsed, raw = parse_excel(excel_path, mapping)

    # 输出
    output = {
        "report_date": datetime.now().strftime("%Y-%m-%d"),
        "source_file": latest,
        "data": parsed,
        "unmapped_count": len(mapping["unmapped_fields"]),
        "unmapped_fields": mapping["unmapped_fields"],
        "notes": mapping["notes"]
    }

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out_path = os.path.join(OUTPUT_DIR, "taobao_parsed.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    # 打印结果
    print(f"\n=== 解析结果 ({len(parsed)} fields) ===")
    for k, v in parsed.items():
        print(f"  {k}: {v}")

    print(f"\nOutput saved to: {out_path}")
    print(f"\n⚠️ 注意: 数据为空（NULL）表示Excel下载失败，需修复RPA脚本")


if __name__ == "__main__":
    main()
